from copy import deepcopy
import os
import numpy as np
import pandas as pd

# For Lasso regression
from numpy import iterable
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline
from sklearn.model_selection import train_test_split, GridSearchCV
from sklearn.linear_model import Lasso
from sklearn.multiclass import OneVsRestClassifier

# For OneHotEncoder
from sklearn.preprocessing import OneHotEncoder
from sklearn.compose import ColumnTransformer
from sklearn.model_selection import GridSearchCV, KFold, cross_val_score

# For RandomForestClassifier
from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier
from sklearn.metrics import confusion_matrix, balanced_accuracy_score
from sklearn.model_selection import StratifiedKFold

# For GroupKFold
from sklearn.preprocessing import LabelEncoder
from sklearn.model_selection import GroupKFold

# For GradientBoost
from sklearn.datasets import make_hastie_10_2
from sklearn.ensemble import GradientBoostingClassifier

# For Report (Confusion Matrix/Classification Report/AUROC)
from sklearn import metrics
from sklearn.metrics import ConfusionMatrixDisplay
from sklearn.svm import SVC
from sklearn.metrics import (
    classification_report,
    f1_score,
    accuracy_score,
    average_precision_score,
)
from sklearn.metrics import roc_auc_score, RocCurveDisplay
from sklearn.model_selection import StratifiedKFold
from sklearn.metrics import roc_curve, auc

# For Decision Tree
from sklearn import tree

from mpl_toolkits.axes_grid1.inset_locator import inset_axes
from joblib import dump, Parallel, delayed
import joblib
from datetime import datetime, timedelta

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
import time
import multiprocessing
import gc



# This module implements training, evaluation, and result export logic for machine learning models (especially random forests, gradient boosting, etc.) within a clinical risk prediction pipeline.
# It includes:

# Cross-validated training with hyperparameter tuning (trained_model)

# Evaluation of ROC/AUPRC performance across train/test/validation sets (eval)

# Export of predicted values and ROC curves for reproducibility and visualization


def parse_n_jobs(n_jobs_indiv):
    """Cleans and computes safe core count based on user input or auto-detects."""
    total_cores = joblib.cpu_count()
    default_safe = max(1, total_cores - 2)

    if n_jobs_indiv is None:
        return default_safe

    if type(n_jobs_indiv) is dict:
        return n_jobs_indiv

    # Handle string input
    if isinstance(n_jobs_indiv, str):
        try:
            n_jobs_indiv = float(n_jobs_indiv)
        except ValueError:
            print(f"⚠️ Warning: Invalid n_jobs_indiv value '{n_jobs_indiv}', defaulting to {default_safe}")
            return default_safe

    # If a float between 0 and 1 → treat as fraction of total cores
    if isinstance(n_jobs_indiv, float) and 0 < n_jobs_indiv <= 1:
        return max(1, int(total_cores * n_jobs_indiv))

    # If an int > 1 → treat as fixed core count, but cap to available
    if isinstance(n_jobs_indiv, (int, float)) and n_jobs_indiv > 1:
        return min(int(n_jobs_indiv), total_cores)

    # Fallback
    print(f"⚠️ Warning: Unsupported n_jobs_indiv value '{n_jobs_indiv}', defaulting to {default_safe}")
    return default_safe





class trained_model:
    """
    Class for training ML models with cross-validation and hyperparameter tuning.

    Attributes
    ----------
    model_with_info : dict
        Stores all trained model objects along with input/output data and evaluation metrics.
    models : list
        List of GridSearchCV objects from each cross-validation fold.
    """

    def __init__(self) -> None:
        pass

    def five_fold_cross_train(
        self,
        self_pip,
        estimator,
        hyperparams,
        grouped_split_on: str = "split_int",
    ):
        """
        Train an estimator using 5-fold cross-validation and grid search, optionally using grouped splitting.

        Parameters
        ----------
        self_pip : Pipeline
            Pipeline object containing data, preprocessing and configuration.
        estimator : sklearn.base.BaseEstimator
            The scikit-learn-compatible estimator to be trained.
        hyperparams : dict
            A dictionary with fixed parameters, grid parameters, and pipeline-level CV config.
        grouped_split_on : str, default="split_int"
            The column used to group samples during outer cross-validation.
        """
        # get the data from the pipeline object:
        X_train, y_train = self_pip.data.X, self_pip.data.y
        models = []
        model_with_info = {}
        iterator = -1

        # get the hyperparameters from the user_input.yaml
        fixed_params = hyperparams["params_fixed"] #Loads the fixed params e.g. specific for "Random Forest Classifier", indexed by "RFC" in user_input.yaml
        grid_params = hyperparams["params_grid"]   #Loads the grid params, with different options that will be assessed in gridsearch
        pipeline_params = hyperparams["params_pipeline"]
        n_splits= pipeline_params["n_splits"]
        random_state = fixed_params["random_state"]

        cv_method = pipeline_params["cross_validation_method"]

        # Set nr of jobs depending on system requirements. n_jobs is limited in windows by max of 60, and additionally limited by your OS hardware.
        # For the two layers, we need to split the budget. With default of 5 folds, we use 5 outer folds (for parallel) and e.g. 3 for the grid search INSIDE of the fold.
        # So if I have 20 cores, we use 18 for modelling, with 5 outer folds we can use 3 (3*15 = 15 and 15 < 18) for the inner folds.
        # Define nr of cores to safely use
        if not (type(pipeline_params.get("n_jobs_indiv")) is dict):
            if int(pipeline_params.get("n_jobs_indiv")) > 1:

                max_cores = multiprocessing.cpu_count()
                n_jobs_indiv = pipeline_params.get("n_jobs_indiv", None) #Get the manual jobs definition from the user_input.yaml. Not recommended, but up to user to use it.

                safe_cores = parse_n_jobs(n_jobs_indiv) #Run the parse function that looks whether the defined n_jobs is a string, int or float and returns the safe cores to use.


                outer_n_jobs = min(n_splits, safe_cores)
                inner_n_jobs = max(1, safe_cores // outer_n_jobs)

                # If the outer and inner jobs exceed the safe cores, adjust the inner jobs
                if outer_n_jobs * inner_n_jobs > safe_cores:
                    inner_n_jobs = max(1, safe_cores // outer_n_jobs)

                print(f"Detected {max_cores} logical cores.")
                print(f"Using {outer_n_jobs} cores for folds (Parallel) and {inner_n_jobs} cores inside GridSearchCV.")
            elif pipeline_params.get("n_jobs_indiv") == -1:
                outer_n_jobs = 1
                inner_n_jobs = -1
            else:
                outer_n_jobs = 1
                inner_n_jobs = 1
                print(f"Using {outer_n_jobs} cores for folds (Parallel) and {inner_n_jobs} cores inside GridSearchCV.")
        else:
            outer_n_jobs = pipeline_params.get("n_jobs_indiv").get("outer_n_jobs", 1)
            inner_n_jobs = pipeline_params.get("n_jobs_indiv").get("inner_n_jobs", 1)
            print(f"Using {outer_n_jobs} cores for folds (Parallel) and {inner_n_jobs} cores inside GridSearchCV.")

        # construct cross-validation method for inner and outer loop for hyperparameter tuning and 5fold training
        if cv_method == "grouped":
            gcv = GroupKFold(n_splits)
            scv = StratifiedKFold(
                n_splits=n_splits,
                shuffle=True,
                random_state=random_state,
            )
            cv_outer, cv_inner = gcv, scv
            le = LabelEncoder()
            cv_out = cv_outer.split(X_train, y_train, groups=le.fit_transform(X_train[grouped_split_on]))

        elif cv_method == "stratified":
            scv = StratifiedKFold(
                n_splits=n_splits,
                shuffle=True,
                random_state=random_state,
            )
            cv_outer, cv_inner = scv, scv
            cv_out = cv_outer.split(X_train, y_train)
            le = LabelEncoder()
        if hyperparams=={}:
            cv_inner = StratifiedKFold(
                n_splits=2,
                shuffle=True,
                random_state=random_state,
            )
        # start outer loop:
        ## define cv object
        def run_inner_loop(
            iterator,
            train_idx,
            test_idx,
            inner_n_jobs,
            estimator=estimator,
            models=models,
            model_with_info=model_with_info,
            self_pip=self_pip
        ):
            X_train_inner, X_test_inner = (
                X_train.iloc[train_idx, :],
                X_train.iloc[test_idx, :],
            )
            if ~self_pip.model_type.startswith("survival"):
                y_train_inner, y_test_inner = (
                    y_train.iloc[train_idx,:],
                    y_train.iloc[test_idx,:],
                )
            else:
                y_train_inner, y_test_inner = (
                    y_train[train_idx],
                    y_train[test_idx],
                )



            # tune the hyperparameters in the inner loop:
            model = GridSearchCV(
                estimator=estimator,
                param_grid=grid_params,
                verbose=pipeline_params["verbose"],
                cv=cv_inner,
                scoring=pipeline_params["scoring_grid_search"],
                n_jobs=inner_n_jobs,
            )  # 'roc_auc_ovr_weighted',‘balanced_accuracy’ also an option look at the overfitting
            if type(y_train_inner) is np.ndarray:
                model.fit(
                    self_pip.ohe.transform(X_train_inner), y_train_inner
                )  # ,groups=le.fit_transform(X_train_inner['split_int']))
            else:
                model.fit(
                    self_pip.ohe.transform(X_train_inner), y_train_inner[self_pip.user_input.target]
                )  # ,groups=le.fit_transform(X_train_inner['split_int']))

            models.append(model)

            if self_pip.model_type.startswith("survival"):
                if hasattr(self_pip.data, "y_val"):
                    score_val = model.score(self_pip.ohe.transform(self_pip.data.X_val), self_pip.data.y_val)
                    pred_val = pd.Series(
                        model.predict_proba(self_pip.ohe.transform(self_pip.data.X_val))[1].tolist(),
                        index=self_pip.data.X_val.index,
                        name=self_pip.user_input.target,
                    )
                else:
                    score_val = None
                    pred_val = None
                model_with_info.update(
                    {
                        f"model_{iterator}": {
                            "model": model,
                            "X_train_inner": X_train_inner,
                            "X_test_inner": X_test_inner,
                            "y_train_inner": y_train_inner,
                            "y_test_inner": y_test_inner,
                            "y_pred_test": pd.Series(
                                model.predict_proba(self_pip.ohe.transform(X_test_inner))[1].tolist(),
                                index=X_test_inner.index,
                                name=self_pip.user_input.target,
                            ),
                            "y_pred_train": pd.Series(
                                model.predict_proba(self_pip.ohe.transform(X_train_inner))[1].tolist(),
                                index=X_train_inner.index,
                                name=self_pip.user_input.target,
                            ),
                            "y_pred_val": pred_val,
                            "score_train": model.score(self_pip.ohe.transform(X_train_inner), y_train_inner),
                            "score_test": model.score(self_pip.ohe.transform(X_test_inner), y_test_inner),
                            "score_val": score_val,
                        }
                    }
                )
            else:

                fpr_test, tpr_test, _ = roc_curve(
                    y_test_inner,
                    model.predict_proba(self_pip.ohe.transform(X_test_inner))[:, 1],
                )
                fpr_train, tpr_train, _ = roc_curve(
                    y_train_inner,
                    model.predict_proba(self_pip.ohe.transform(X_train_inner))[:, 1],
                )
                model_with_info.update(
                    {
                        f"model_{iterator}": {
                            "model": model,
                            "X_train_inner": X_train_inner,
                            "X_test_inner": X_test_inner,
                            "y_train_inner": y_train_inner,
                            "y_test_inner": y_test_inner,
                            "y_pred_test": pd.Series(
                                model.predict_proba(self_pip.ohe.transform(X_test_inner))[:, 1],
                                index=X_test_inner.index,
                                name=self_pip.user_input.target,
                            ),
                            "y_pred_train": pd.Series(
                                model.predict_proba(self_pip.ohe.transform(X_train_inner))[:, 1],
                                index=X_train_inner.index,
                                name=self_pip.user_input.target,
                            ),
                            "perf_on_test": {
                                "fpr": fpr_test,
                                "tpr": tpr_test,
                                "roc_auc_raw": auc(fpr_test, tpr_test),
                            },
                            "perf_on_train": {
                                "fpr": fpr_train,
                                "tpr": tpr_train,
                                "roc_auc_raw": auc(fpr_train, tpr_train),
                            },
                            "user_input": self_pip.user_input,
                        }
                    }
                )
                gc.collect()
            return model_with_info, models

        with Parallel(n_jobs=outer_n_jobs,backend='loky') as parallel: ## TODO: running into issues here -> so only using one core but tuning  the hypereparamerter seach ## or 'loky'
            results = parallel(
            delayed(run_inner_loop)(iterator, train_idx, test_idx, inner_n_jobs)
            for iterator, (train_idx, test_idx) in enumerate(cv_out)
            )
        for model_with_i, model in results:
            model_with_info.update(model_with_i)
            models.append(model)
        self.model_with_info = model_with_info
        self.models = models

    def roc_test_train(self, key_df, ax=None, interpolate=False):
        tprs = pd.DataFrame()
        aucs = []
        for key, val in self.model_with_info.items():
            if key.startswith("model"):
                fpr = val.get(key_df).get("fpr")
                tpr = val.get(key_df).get("tpr")
                fpr_base = np.linspace(0, 1, 100)
                tpr_interpol = np.interp(fpr_base, fpr, tpr)
                auc_interpol = auc(fpr_base, tpr_interpol)
                if interpolate:
                    ax.plot(fpr_base, tpr_interpol, color="lightgray", lw=2)
                else:
                    ax.plot(fpr, tpr, color="lightgray", lw=2)
                tprs[key_df + "_interpol_tpr_" + key] = tpr_interpol
                aucs.append(auc_interpol)
        if interpolate:
            tprs["mean_tpr"] = tprs.mean(axis=1)
            ax.plot(
                fpr_base,
                tprs.mean_tpr,
                lw=2,
                label=f"Mean ROCcurve(AUC = {round(np.mean(aucs),2)})",
            )
        ax.legend()
        ax.set_title(key_df)
        self.model_with_info.update({key_df: {"tprs_interpol": tprs, "aucs_interpol": aucs}})


class eval:
    def __init__(self, pip_self, only_val=False):
        self.only_val = only_val
        """Evaluate the performance of the models on the training and testing sets, as well as the performance of the mastermodel on the validation model.


        Args:
            pip_self (pipeline_object): pipeline object of the model
            only_val (bool, optional): If you are only interested on the performance on the validation set you can set is to True (not recommened). Defaults to False.
        """

        if not only_val:
            # generate some print outputs:
            test_, train_ = [], []
            print("\n")
            print("Evaluation of the Model".center(80, "-"))
            for key, val in pip_self.trained_model.model_with_info.items():
                print(key.center(80, "-"))
                test = val["perf_on_test"]["roc_auc_raw"]
                print("auc_test:", test)
                train = val["perf_on_train"]["roc_auc_raw"]
                print("auc_train:", train)
                print("best_params:", val["model"].best_params_)
                print("".center(80, "-"))
                test_.append(test)
                train_.append(train)
            print("ROCcurve AUC".center(80, "-"))
            print("test_mean:", np.mean(test_))
            print("train_mean:", np.mean(train_))

        ### now get the TPRS and AUCS
        tprs_test = pd.DataFrame()
        tprs_train = pd.DataFrame()
        tprs_val = pd.DataFrame()
        aucs_test = []
        aucs_train = []
        aucs_val = []
        auprcs_test = []
        auprcs_train = []
        auprcs_val = []
        ohe = pip_self.ohe

        has_validation_data = (
            hasattr(pip_self.data, "y_val") and pip_self.data.y_val is not None
        )  # Check whether validation data exists
        print(f"The training and testing is evaluated on the {pip_self.user_input.target} labels.")
        print(
            f"The y_val is now set to the target to validate on: {pip_self.user_input.target_to_validate_on}\nFrom now on this will be used as y_val! change if needed."
        )
        print(f"Using {pip_self.user_input.target_to_validate_on} for the evaluation on the validation set.")

        for key, val in pip_self.trained_model.model_with_info.items():
            model = val["model"]
            if not only_val:
                # for the testing:
                fpr, tpr, thres = roc_curve(val["y_test_inner"], val["y_pred_test"].tolist())
                fpr_test, tpr_test, thres_test = fpr.copy(), tpr.copy(), thres.copy()
                fpr_base = np.linspace(0, 1, 100)
                tpr = np.interp(fpr_base, fpr, tpr)
                aucs_test.append(auc(fpr_base, tpr))
                tprs_test[key] = tpr
                auprcs_test.append(average_precision_score(val["y_test_inner"], val["y_pred_test"].tolist()))
                # for the training
                fpr, tpr, thres = roc_curve(val["y_train_inner"], val["y_pred_train"].tolist())
                fpr_train, tpr_train, thres_train = fpr.copy(), tpr.copy(), thres.copy()
                fpr_base = np.linspace(0, 1, 100)
                tpr = np.interp(fpr_base, fpr, tpr)
                aucs_train.append(auc(fpr_base, tpr))
                tprs_train[key] = tpr
                auprcs_train.append(average_precision_score(val["y_train_inner"], val["y_pred_train"].tolist()))

            # for the validation
            if has_validation_data:
                # set the y_val to the desired target
                pip_self.data.y_val = pip_self.data.y_val_orig[pip_self.user_input.target_to_validate_on].copy()

                val.update(
                    {
                        "y_val": pip_self.data.y_val_orig[pip_self.user_input.target_to_validate_on],
                        "y_pred_val": pd.Series(
                            model.predict_proba(ohe.transform(pip_self.data.X_val))[:, 1],
                            index=pip_self.data.X_val.index,
                            name=pip_self.user_input.target_to_validate_on,
                        ),
                    }
                )
                fpr, tpr, thres = roc_curve(val.get("y_val"), val.get("y_pred_val").to_list())
                fpr_val, tpr_val, thres_val = fpr.copy(), tpr.copy(), thres.copy()
                fpr_base = np.linspace(0, 1, 100)
                tpr = np.interp(fpr_base, fpr, tpr)
                aucs_val.append(auc(fpr_base, tpr))
                tprs_val[key] = tpr
                auprcs_val.append(average_precision_score(val.get("y_val"), val.get("y_pred_val").to_list()))
            else:
                print("Validation data not available")
        if not only_val:
            self.train = {
                "tprs": tprs_train,
                "aucs": aucs_train,
                "auprcs": auprcs_train,
                "raw": [fpr_train, tpr_train, thres_train],
            }
            self.test = {
                "tprs": tprs_test,
                "aucs": aucs_test,
                "auprcs": auprcs_test,
                "raw": [fpr_test, tpr_test, thres_test],
            }
        if has_validation_data:
            self.val = {
                "tprs": tprs_val,
                "aucs": aucs_val,
                "auprcs": auprcs_val,
                "raw": [fpr_val, tpr_val, thres_val],
            }
        else:
            self.val = {}
        self.feature_imp = {"PlotX": None}
        self.test_train_pred = {}


    def save_performance_combination(self, pip_self, tprs, pred_values, y_true,
                                    true_cancerreg=None, cohort=None, save_format="joblib"):
        """
        Saves the performance metrics (TPRS and prediction values) of a model run.

        Supports saving in .joblib (default), .csv, or .xlsx formats.

        Args:
            pip_self: Pipeline object containing user input and paths.
            tprs (pd.DataFrame): TPRS data from the model run.
            pred_values (pd.Series or pd.DataFrame): Predicted values from the model run.
            y_true (pd.Series or np.array): True labels.
            true_cancerreg (optional): Placeholder for cancer regression metadata.
            cohort (str, optional): Cohort label ('train', 'val', 'test').
            save_format (str, optional): File format to save data in: 'joblib', 'csv', or 'xlsx'.

        Returns:
            None
        """
        print(f"Starting save_performance_combination using format: {save_format}")
        #Put together identifier (column_name in combined_output) based on row/column subset and optional suffix
        row_subset = pip_self.user_input.row_subset
        col_subset = pip_self.user_input.col_subset

        identifier = f"{row_subset}_{col_subset}"
        if hasattr(pip_self.user_input, 'pl_suffix') and pip_self.user_input.pl_suffix is not None:
            identifier = f"{row_subset}_{col_subset}" + f"_{pip_self.user_input.pl_suffix}"
        estimator = pip_self.model_type

        combined_output_path = os.path.join(
            pip_self.pipeline_output_path,
            "combined_output" if cohort is None else f"combined_output/{cohort}"
        )
        os.makedirs(combined_output_path, exist_ok=True)

        # ---------- Save TPRS ----------
        tprs_export = tprs.rename(columns=lambda x: f"{identifier}_{str(x)}")

        if save_format == "joblib":
            tprs_combined_path = os.path.join(combined_output_path, "TPRS_combined.joblib")
            if os.path.exists(tprs_combined_path):
                tprs_combined = joblib.load(tprs_combined_path)
            else:
                tprs_combined = pd.DataFrame()

            tprs_combined = tprs_combined.drop(columns=[col for col in tprs_combined.columns
                                                        if col.startswith(f"{identifier}_")], errors='ignore')
            tprs_combined = pd.concat([tprs_combined, tprs_export], axis=1)
            joblib.dump(tprs_combined, tprs_combined_path)
            print(f"TPRS data saved to {tprs_combined_path}")

        else:
            tprs_combined_path = os.path.join(combined_output_path, "TPRS_combined.xlsx")
            if os.path.exists(tprs_combined_path):
                tprs_combined = pd.read_excel(tprs_combined_path)
            else:
                tprs_combined = pd.DataFrame()

            current_cols = [col for col in tprs_combined.columns if col.startswith(f"{identifier}_")]
            if current_cols:
                tprs_combined = tprs_combined.drop(columns=current_cols)

            tprs_combined_export = pd.concat([tprs_combined, tprs_export], axis=1)
            if save_format == "csv":
                tprs_combined_export.to_csv(tprs_combined_path.replace(".xlsx", ".csv"), index=False)
            else:
                tprs_combined_export.to_excel(tprs_combined_path, index=False)
            print(f"TPRS data exported to {tprs_combined_path}")

        # ---------- Save Prediction Values ----------
        if cohort in ["val", "train", "test"]:
            pred_df = pip_self.eval.test_train_pred.get(cohort)     #Mean Prediction values for the cohort
            if cohort == "val": #
                for model, value in pip_self.trained_model.model_with_info.items(): #Loop over 5 cross validation models to store all respective prediction values
                    pred_df[f"y_pred_{cohort}_" + model] = value.get(f"y_pred_{cohort}") #Only for val, for test/train this takes up too much memory


            if pred_df is None:
                print(f"Skipping prediction values export: no data found for cohort '{cohort}'")
                return

            pred_df = pred_df.reset_index()
            sheet_name = identifier

            if save_format == "joblib":
                prediction_combined_path = os.path.join(combined_output_path, "Prediction_values_combined.joblib")
                if os.path.exists(prediction_combined_path):
                    pred_combined = joblib.load(prediction_combined_path)
                else:
                    pred_combined = {}

                pred_combined[identifier] = pred_df
                joblib.dump(pred_combined, prediction_combined_path)
                print(f"Prediction values saved to {prediction_combined_path}")

            else:
                prediction_combined_path = os.path.join(combined_output_path, "Prediction_values_combined.xlsx")
                if not os.path.exists(prediction_combined_path):
                    wb = Workbook()
                    wb.save(prediction_combined_path)
                    wb.close()
                    time.sleep(0.1)

                wb = load_workbook(prediction_combined_path)
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]
                wb.save(prediction_combined_path)
                wb.close()
                time.sleep(0.1)

                with pd.ExcelWriter(prediction_combined_path, mode="a", engine="openpyxl") as writer:
                    pred_df.to_excel(writer, sheet_name=sheet_name, index=False)
                print(f"Prediction values exported to {prediction_combined_path}")

        else:
            print("Skipping prediction values export (not a valid cohort or missing data)")


    def get_pv_test_train(self, pip_self):
        """Get the predicted values for the training, testing and validation datasets.
        The predicted values are stored in seperated dataframes that are exported and in the eval.val, eval.test and eval.train dictionaries.


        Args:
            pip_self (_type_): _description_

        Returns:
            _type_: _description_
        """

        export_test = pd.DataFrame()
        export_train = pd.DataFrame()
        export_val = None

        if not self.only_val:
            for key, val in pip_self.trained_model.model_with_info.items():
                y_true_train, y_true_test = [], []
                y_pred_train, y_pred_test = [], []
                eid_train, eid_test = [], []

                # iter through all models and get the performance
                y_true_train = y_true_train + list(val["y_train_inner"][pip_self.user_input.target])
                y_pred_train = y_pred_train + list(val["y_pred_train"])
                eid_train = eid_train + list(val["y_train_inner"].index)
                y_true_test = y_true_test + list(val["y_test_inner"][pip_self.user_input.target])
                y_pred_test = y_pred_test + list(val["y_pred_test"])
                eid_test = eid_test + list(val["y_test_inner"].index)

                # export the stats:
                export_train = pd.concat(
                    [
                        export_train,
                        pd.DataFrame(
                            {
                                "y_true": y_true_train,
                                "y_pred": y_pred_train,
                                "eid": eid_train,
                            }
                        ),
                    ]
                )

                export_test = pd.concat(
                    [
                        export_test,
                        pd.DataFrame(
                            {
                                "y_true": y_true_test,
                                "y_pred": y_pred_test,
                                "eid": eid_test,
                            }
                        ),
                    ]
                )
            export_train.set_index("eid")
            export_test.set_index("eid")

        if hasattr(pip_self.data, "y_val") and pip_self.data.y_val is not None:
            try:
                export_val = deepcopy(pip_self.data.y_val_orig)
                export_val["y_pred"] = pip_self.master_RFC.predict_proba(
                    pip_self.ohe.transform(pip_self.data.X_val)
                ).tolist()
                self.val.update({"predicted_values": export_val})
            except Exception as e:
                print(f"Could not calculate the prediction values for the validation dataset: {e}")
                export_val = None
        else:
            print("Validation dataset not available.")

        if not self.only_val:
            # merge with the y_orig or y_val_orig to get the additional data from the cancerreg...
            export_train = export_train.merge(pip_self.data.y_orig, left_on="eid", right_index=True, how="left")
            export_test = export_test.merge(pip_self.data.y_orig, left_on="eid", right_index=True, how="left")

            self.test_train_pred.update({"train": export_train, "test": export_test, "val": export_val})
            self.train.update({"predicted_values": export_train})
            self.test.update({"predicted_values": export_test})

        self.val.update({"predicted_values": export_val})
        self.test_train_pred.update({"val": export_val})

        if not hasattr(self, "val") or self.val is None:
            self.val = {}
        self.val["predicted_values"] = export_val

        return export_train, export_test, export_val

    def get_metrics(y_true, y_pred, threshold_steps=0.01):
        """A computational expensive function for the evaluation of a model without taking one specific threshold

        Args:
            y_true (_type_): _description_
            y_pred (_type_): _description_
            threshold_steps (float, optional): _description_. Defaults to 0.01.
        """
